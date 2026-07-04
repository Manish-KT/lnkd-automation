"""Excel tracker for outreach state. One row per person ever touched."""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

COLUMNS = [
    "Timestamp",
    "Company",
    "Company URL",
    "Name",
    "Title",
    "Profile URL",
    "Status",          # pending | sent | skipped | failed | already_connected
    "Note Sent",
    "Keyword",
    "Error",
]

STATUS_FILL = {
    "sent": PatternFill("solid", fgColor="C6EFCE"),
    "failed": PatternFill("solid", fgColor="FFC7CE"),
    "skipped": PatternFill("solid", fgColor="FFEB9C"),
    "already_connected": PatternFill("solid", fgColor="D9D9D9"),
}


@dataclass
class Contact:
    company: str
    company_url: str
    name: str
    title: str
    profile_url: str
    keyword: str
    status: str = "pending"
    note_sent: str = ""
    error: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


class Tracker:
    def __init__(self, path: str):
        self.path = path
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if os.path.exists(path):
            self.wb = load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Outreach"
            self.ws.append(COLUMNS)
            for cell in self.ws[1]:
                cell.font = Font(bold=True)
            self._save()
        self._seen_urls = self._load_seen_urls()
        self._processed_companies = self._load_processed_companies()

    def _load_seen_urls(self) -> set[str]:
        urls: set[str] = set()
        url_col = COLUMNS.index("Profile URL") + 1
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row and len(row) >= url_col and row[url_col - 1]:
                urls.add(self._normalize_url(str(row[url_col - 1])))
        return urls
    
    def _load_processed_companies(self) -> set[str]:
        company_col = COLUMNS.index("Company") + 1
        companies: set[str] = set()
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if row and row[company_col - 1]:
                companies.add(
                    str(row[company_col - 1]).strip().lower()
                )
        return companies

    @staticmethod
    def _normalize_url(url: str) -> str:
        return url.split("?")[0].rstrip("/").lower()

    def already_contacted(self, profile_url: str) -> bool:
        return self._normalize_url(profile_url) in self._seen_urls

    def sent_today(self) -> int:
        today = datetime.now().date().isoformat()
        ts_col = COLUMNS.index("Timestamp") + 1
        status_col = COLUMNS.index("Status") + 1
        count = 0
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            ts = str(row[ts_col - 1] or "")
            status = str(row[status_col - 1] or "")
            if ts.startswith(today) and status == "sent":
                count += 1
        return count

    def record(self, contact: Contact) -> None:
        self.ws.append([
            contact.timestamp,
            contact.company,
            contact.company_url,
            contact.name,
            contact.title,
            contact.profile_url,
            contact.status,
            contact.note_sent,
            contact.keyword,
            contact.error,
        ])
        fill = STATUS_FILL.get(contact.status)
        if fill:
            self.ws.cell(row=self.ws.max_row, column=COLUMNS.index("Status") + 1).fill = fill
        self._seen_urls.add(self._normalize_url(contact.profile_url))
        self._processed_companies.add(contact.company.strip().lower())
        self._save()

    def _save(self) -> None:
        # Auto-size columns roughly to content width.
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            self.ws.column_dimensions[self.ws.cell(row=1, column=col_idx).column_letter].width = max(
                len(col_name) + 2, 18
            )
        self.wb.save(self.path)
    
    def get_processed_companies(self) -> set[str]:
        return self._processed_companies