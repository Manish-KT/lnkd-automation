"""Supabase + Excel tracker for outreach state. One row per person ever touched."""
from __future__ import annotations

import os
from dataclasses import dataclass, field, asdict
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from supabase import create_client, Client

from dotenv import load_dotenv
load_dotenv()  # Load environment variables from .env file

# --- Excel Configuration ---
COLUMNS = [
    "Timestamp",
    "Company",
    "Company URL",
    "Name",
    "Title",
    "Profile URL",
    "Status",
    "Note Sent",
    "Keyword",
    "Error",
]

STATUS_FILL = {
    "sent": PatternFill("solid", fgColor="C6EFCE"),
    "sent_without_note": PatternFill("solid", fgColor="C6EFCE"),
    "failed": PatternFill("solid", fgColor="FFC7CE"),
    "failed_error_msg": PatternFill("solid", fgColor="FFC7CE"),
    "skipped": PatternFill("solid", fgColor="FFEB9C"),
    "already_connected": PatternFill("solid", fgColor="D9D9D9"),
}

@dataclass
class Contact:
    company: str
    company_url: str
    name: str
    title: str
    profile_url: str
    keyword: str
    status: str = "pending"
    note_sent: str = ""
    error: str = ""
    timestamp: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))


class Tracker:
    def __init__(self, path: str):
        self.path = path
        
        # ==========================================
        # 1. Initialize Supabase
        # ==========================================
        supabase_url = os.getenv("SUPABASE_URL")
        supabase_key = os.getenv("SUPABASE_KEY")
        
        if not supabase_url or not supabase_key:
            raise ValueError(
                "Missing credentials. Please set SUPABASE_URL and SUPABASE_KEY environment variables."
            )
            
        self.supabase: Client = create_client(supabase_url, supabase_key)
        self.table_name = "outreach"
        
        # ==========================================
        # 2. Initialize Local Excel File
        # ==========================================
        os.makedirs(os.path.dirname(path), exist_ok=True)
        if os.path.exists(path):
            self.wb = load_workbook(path)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "Outreach"
            self.ws.append(COLUMNS)
            
            # Format headers and auto-size columns once during creation
            for col_idx, cell in enumerate(self.ws[1], start=1):
                cell.font = Font(bold=True)
                col_name = COLUMNS[col_idx - 1]
                self.ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 2, 18)
            self.save()

        # ==========================================
        # 3. Hydrate Caches
        # ==========================================
        self._seen_urls = set()
        self._processed_companies = set()
        
        # Hydrate local cache on startup from Supabase (Source of Truth)
        self._load_initial_data()

    def _load_initial_data(self) -> None:
        """
        Loads existing data from Supabase into memory for fast O(1) lookups.
        Handles Supabase's default 1000-row pagination limit.
        """
        limit = 1000
        offset = 0
        
        while True:
            response = (
                self.supabase.table(self.table_name)
                .select("profile_url, company")
                .range(offset, offset + limit - 1)
                .execute()
            )
            
            data = response.data
            if not data:
                break
            
            for row in data:
                if row.get("profile_url"):
                    self._seen_urls.add(self._normalize_url(row["profile_url"]))
                if row.get("company"):
                    self._processed_companies.add(str(row["company"]).strip().lower())
            
            if len(data) < limit:
                break
            
            offset += limit

    @staticmethod
    def _normalize_url(url: str) -> str:
        return url.split("?")[0].rstrip("/").lower()

    def already_contacted(self, profile_url: str) -> bool:
        return self._normalize_url(profile_url) in self._seen_urls

    def sent_today(self) -> int:
        """Queries Supabase directly to accurately count today's successful requests."""
        today_start = datetime.now().date().isoformat()
        
        response = (
            self.supabase.table(self.table_name)
            .select("id", count="exact")
            .gte("timestamp", today_start)
            .in_("status", ["sent", "sent_without_note"])
            .execute()
        )
            
        return response.count if response.count is not None else 0

    def record(self, contact: Contact) -> None:
        """Inserts a record into Supabase AND appends to the local Excel file in memory."""
        # 1. Insert into Supabase DB
        data = asdict(contact)
        self.supabase.table(self.table_name).insert(data).execute()
        
        # 2. Append to Local Excel Workbook
        self.ws.append([
            contact.timestamp,
            contact.company,
            contact.company_url,
            contact.name,
            contact.title,
            contact.profile_url,
            contact.status,
            contact.note_sent,
            contact.keyword,
            contact.error,
        ])
        
        # Apply Excel formatting based on status
        fill = STATUS_FILL.get(contact.status)
        if fill:
            self.ws.cell(row=self.ws.max_row, column=COLUMNS.index("Status") + 1).fill = fill
        
        # 3. Update local sets to prevent duplicates in the current run
        self._seen_urls.add(self._normalize_url(contact.profile_url))
        self._processed_companies.add(contact.company.strip().lower())

    def save(self) -> None:
        """Persists the local Excel workbook to disk."""
        self.wb.save(self.path)
    
    def get_processed_companies(self) -> set[str]:
        return self._processed_companies